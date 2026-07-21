from flask import request, redirect, url_for, flash
from app.admin import bp
from app.admin.core.auth import login_required
from app.admin.core.logger import log_admin_action
from app.db import get_db_connection

@bp.route('/contact_settings/update', methods=['POST'])
@login_required
def update_contact_settings():
    """
    Обновление контактной информации организации (телефон, email, адрес и т.д.).
    Данные сохраняются в таблице contact_settings.
    """
        
    title = request.form.get('title')
    org_name = request.form.get('org_name')
    phones = request.form.get('phones')
    email = request.form.get('email')
    address = request.form.get('address')
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE contact_settings
        SET title = ?, org_name = ?, phones = ?, email = ?, address = ?
        WHERE id = 1
    ''', (title, org_name, phones, email, address))
    conn.commit()
    
    log_admin_action('UPDATE', 'contacts', details=f'Обновлены контактные данные организации')
    
    flash('Настройки контактов успешно обновлены!', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/contact_requests/<int:id>/mark_read', methods=['POST'])
@login_required
def mark_request_read(id):
    """
    Пометка входящей заявки с формы обратной связи как "прочитанной".
    """
        
    conn = get_db_connection()
    conn.execute('UPDATE contact_requests SET status = "read" WHERE id = ?', (id,))
    conn.commit()
    
    log_admin_action('UPDATE', 'contacts', entity_id=id, details=f'Входящая заявка ID {id} отмечена как прочитанная')
    
    flash('Заявка отмечена как прочитанная.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/contact_requests/<int:id>/delete', methods=['POST'])
@login_required
def delete_request(id):
    """
    Удаление входящей заявки (формы обратной связи) из базы данных.
    """
        
    conn = get_db_connection()
    
    req = conn.execute('SELECT name FROM contact_requests WHERE id = ?', (id,)).fetchone()
    req_name = req['name'] if req else f"ID {id}"
    
    conn.execute('DELETE FROM contact_requests WHERE id = ?', (id,))
    conn.commit()
    
    log_admin_action('DELETE', 'contacts', entity_id=id, details=f'Удалена входящая заявка от "{req_name}"')
    
    flash('Заявка удалена.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/delete_submission/<int:id>', methods=['POST'])
@login_required
def delete_submission(id):
    """
    Удаление конкурсной заявки (динамической формы) из базы данных.
    """
        
    conn = get_db_connection()
    conn.execute('DELETE FROM form_submissions WHERE id = ?', (id,))
    conn.commit()
    
    log_admin_action('DELETE', 'forms_data', entity_id=id, details=f'Удалена конкурсная заявка ID {id}')
    
    flash('Заявка удалена.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/export_submissions/<int:form_id>')
@login_required
def export_submissions(form_id):
    """
    Выгрузка заявок конкретного конкурса в Excel.
    """
        
    from flask import send_file
    import io
    import json
    import openpyxl
    from openpyxl.utils import get_column_letter

    conn = get_db_connection()
    form = conn.execute('SELECT * FROM page_forms WHERE id = ?', (form_id,)).fetchone()
    if not form:
        flash('Форма не найдена.', 'danger')
        return redirect(url_for('admin.dashboard', tab='contacts'))
        
    submissions = conn.execute('SELECT * FROM form_submissions WHERE form_id = ? ORDER BY created_at ASC', (form_id,)).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    if not submissions:
        ws.append(["Нет данных"])
    else:
        # Получаем все возможные колонки
        columns = []
        parsed_submissions = []
        for s in submissions:
            data = json.loads(s['submission_data'])
            parsed_submissions.append((s['created_at'], data))
            for k in data.keys():
                if k not in columns:
                    columns.append(k)
                    
        headers = ["Дата подачи"] + columns
        ws.append(headers)
        
        # Стилизуем заголовок
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 25
            
        for created_at, data in parsed_submissions:
            row = [created_at]
            for col in columns:
                row.append(data.get(col, ""))
            ws.append(row)
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    # Формируем безопасное имя файла
    filename = f"export_form_{form_id}.xlsx"
    
    log_admin_action('UPLOAD', 'forms_data', entity_id=form_id, details=f'Выгружены заявки формы ID {form_id}')
    
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
