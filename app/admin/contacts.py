from flask import request, redirect, url_for, flash, session
from app.admin import bp
from app.db import get_db_connection

@bp.route('/contact_settings/update', methods=['POST'])
def update_contact_settings():
    """
    Обновление контактной информации организации (телефон, email, адрес и т.д.).
    Данные сохраняются в таблице contact_settings.
    """
    if not session.get('is_admin'):
        return redirect(url_for('admin.login'))
        
    title = request.form.get('title')
    org_name = request.form.get('org_name')
    phones = request.form.get('phones')
    email = request.form.get('email')
    address = request.form.get('address')
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE contact_settings
        SET title = ?, org_name = ?, phones = ?, email = ?, address = ?
        WHERE id = 1
    ''', (title, org_name, phones, email, address))
    conn.commit()
    conn.close()
    
    flash('Настройки контактов успешно обновлены!', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/contact_requests/<int:id>/mark_read', methods=['POST'])
def mark_request_read(id):
    """
    Пометка входящей заявки с формы обратной связи как "прочитанной".
    """
    if not session.get('is_admin'):
        return redirect(url_for('admin.login'))
        
    conn = get_db_connection()
    conn.execute('UPDATE contact_requests SET status = "read" WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    flash('Заявка отмечена как прочитанная.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/contact_requests/<int:id>/delete', methods=['POST'])
def delete_request(id):
    """
    Удаление входящей заявки (формы обратной связи) из базы данных.
    """
    if not session.get('is_admin'):
        return redirect(url_for('admin.login'))
        
    conn = get_db_connection()
    conn.execute('DELETE FROM contact_requests WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    flash('Заявка удалена.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/delete_submission/<int:id>', methods=['POST'])
def delete_submission(id):
    """
    Удаление конкурсной заявки (динамической формы) из базы данных.
    """
    if not session.get('is_admin'):
        return redirect(url_for('admin.login'))
        
    conn = get_db_connection()
    conn.execute('DELETE FROM form_submissions WHERE id = ?', (id,))
    conn.commit()
    conn.close()
    
    flash('Заявка удалена.', 'success')
    return redirect(url_for('admin.dashboard', tab='contacts'))

@bp.route('/export_submissions/<int:form_id>')
def export_submissions(form_id):
    """
    Выгрузка заявок конкретного конкурса в Excel.
    """
    if not session.get('is_admin'):
        return redirect(url_for('admin.login'))
        
    from flask import send_file
    import io
    import json
    import openpyxl
    from openpyxl.utils import get_column_letter

    conn = get_db_connection()
    form = conn.execute('SELECT * FROM page_forms WHERE id = ?', (form_id,)).fetchone()
    if not form:
        conn.close()
        flash('Форма не найдена.', 'danger')
        return redirect(url_for('admin.dashboard', tab='contacts'))
        
    submissions = conn.execute('SELECT * FROM form_submissions WHERE form_id = ? ORDER BY created_at ASC', (form_id,)).fetchall()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"
    
    if not submissions:
        ws.append(["Нет данных"])
    else:
        # Получаем все возможные колонки
        columns = []
        parsed_submissions = []
        for s in submissions:
            data = json.loads(s['submission_data'])
            parsed_submissions.append((s['created_at'], data))
            for k in data.keys():
                if k not in columns:
                    columns.append(k)
                    
        headers = ["Дата подачи"] + columns
        ws.append(headers)
        
        # Стилизуем заголовок
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = openpyxl.styles.Font(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 25
            
        for created_at, data in parsed_submissions:
            row = [created_at]
            for col in columns:
                row.append(data.get(col, ""))
            ws.append(row)
            
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    # Формируем безопасное имя файла
    filename = f"export_form_{form_id}.xlsx"
    return send_file(
        out,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
